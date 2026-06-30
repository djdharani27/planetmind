import csv
import io
import json
from pathlib import Path
from backend.logging_config import logger


def extract_csv_text(file_path: Path) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = [", ".join(row) for row in reader]
            return "\n".join(rows)
    except Exception as e:
        logger.error(f"CSV extraction failed: {e}")
        return ""


def extract_xlsx_text(file_path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                parts.append("\t".join([str(cell) if cell is not None else "" for cell in row]))
        wb.close()
        return "\n".join(parts)
    except Exception as e:
        logger.error(f"XLSX extraction failed: {e}")
        return ""


def extract_eml_text(file_path: Path) -> str:
    try:
        import email
        from email import policy
        with open(file_path, "rb") as f:
            msg = email.message_from_binary_file(f, policy=policy.default)
        parts = []
        if msg["subject"]:
            parts.append(f"Subject: {msg['subject']}")
        if msg["from"]:
            parts.append(f"From: {msg['from']}")
        if msg["to"]:
            parts.append(f"To: {msg['to']}")
        if msg["date"]:
            parts.append(f"Date: {msg['date']}")
        parts.append("")
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    parts.append(payload.decode("utf-8", errors="replace"))
        return "\n".join(parts)
    except Exception as e:
        logger.error(f"EML extraction failed: {e}")
        return ""


def extract_raw_text(file_path: Path) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as e:
        logger.error(f"Raw text extraction failed: {e}")
        return ""


def extract_text_by_type(file_path: Path, file_type: str) -> str:
    if file_type == "text/csv":
        return extract_csv_text(file_path)
    elif file_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
        return extract_xlsx_text(file_path)
    elif file_type in ("message/rfc822", "application/vnd.ms-outlook"):
        return extract_eml_text(file_path)
    elif file_type == "text/plain":
        return extract_raw_text(file_path)
    return ""
